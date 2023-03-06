import os
import openpyxl


def Create_excel(excel_name):
    wb = openpyxl.Workbook()
    # wb = openpyxl.load_workbook('./Vit_record.xlsx')
    sheet = wb.active
    sheet.title = 'Metrics'

    # wb.sheetnames

    # sheet = wb['ViT']
    sheet.cell(row=1, column=2).value = 'Accuracy'
    sheet.cell(row=1, column=3).value = 'Recall'
    sheet.cell(row=1, column=4).value = 'Precision'
    sheet.cell(row=1, column=5).value = 'F1-score'
    sheet.cell(row=1, column=6).value = 'FDR'
    sheet.cell(row=1, column=7).value = 'Error Rate'
    print(sheet.cell(row=2, column=2).value)
    # sheet = wb.active
    # sheet.title = 'Test'

    # wb.sheetnames

    wb.save(excel_name)


file_name = '/media/hsy/1882C80C82C7EC76/ZJH/Static/confu/'
file_path = [file_name + i for i in os.listdir(file_name)]
excel_name = '/media/hsy/1882C80C82C7EC76/ZJH/Static/Metrics.xlsx'
if not os.path.exists(excel_name):
    Create_excel(excel_name)
sawb = openpyxl.load_workbook(excel_name)
save_sheet = sawb.active

row_idx = 0
for file in file_path:
    actual_row = 2+row_idx*5
    save_sheet.cell(row=actual_row, column=1).value = file[45:-5]
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    for k in range(5):
        T1 = sheet.cell(row=k*3+1, column=1).value
        F12 = sheet.cell(row=k*3+1, column=2).value
        F13 = sheet.cell(row=k*3+1, column=3).value
        T2 = sheet.cell(row=1+k*3+1, column=2).value
        F21 = sheet.cell(row=1+k*3+1, column=1).value
        F23 = sheet.cell(row=1+k*3+1, column=3).value
        T3 = sheet.cell(row=2+k*3+1, column=3).value
        F31 = sheet.cell(row=2+k*3+1, column=1).value
        F32 = sheet.cell(row=2+k*3+1, column=2).value

        Accuracy = (T1 + T2 + T3) / (T1+T2+T3+F12+F13+F21+F23+F31+F32)
        Error_rate = 1 - Accuracy

        Recall_1 = T1 / (T1 + F12 + F13)
        Precision_1 = T1 / (T1 + F21 + F31)
        FDR_1 = (F12 + F31) / (T1 + F12 + F13)
        F1_S_1 = 2 * Recall_1 * Precision_1 / (Precision_1 + Recall_1)

        Recall_2 = T2 / (T2 + F21 + F23)
        Precision_2 = T2 / (T2 + F12 + F32)
        FDR_2 = (F21 + F23) / (T2 + F21 + F23)
        F1_S_2 = 2 * Recall_2 * Precision_2 / (Precision_2 + Recall_2)

        Recall_3 = T3 / (T3 + F31 + F32)
        Precision_3 = T3 / (T3 + F13 + F23)
        FDR_3 = (F32 + F31) / (T3 + F31 + F32)
        F1_S_3 = 2 * Recall_3 * Precision_3 / (Precision_3 + Recall_3)

        Recall = (Recall_1 + Recall_2 + Recall_3) / 3
        Precision = (Precision_1 + Precision_2 + Precision_3) / 3
        FDR = (FDR_1 + FDR_2 + FDR_3) / 3
        F1_S = (F1_S_1 + F1_S_2 + F1_S_3) / 3

        save_sheet.cell(row=2+row_idx*5+k, column=2).value = Accuracy
        save_sheet.cell(row=2+row_idx*5+k, column=3).value = Recall
        save_sheet.cell(row=2+row_idx*5+k, column=4).value = Precision
        save_sheet.cell(row=2+row_idx*5+k, column=5).value = F1_S
        save_sheet.cell(row=2+row_idx*5+k, column=6).value = FDR
        save_sheet.cell(row=2+row_idx*5+k, column=7).value = Error_rate
    row_idx += 1
sawb.save(excel_name)

# f = open(file_path, 'r')
# file = f.read()
# search = 'Confusion Matrix:'

# print(f)

