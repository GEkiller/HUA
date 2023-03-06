import os
import json

import torch
from PIL import Image
from torchvision import transforms
from collections import Counter

from My_model import volo_dense
# from models.volo import *
from my_utils import read_csv_test
import time
import openpyxl
from excel_test import Create_excel
from assist_block import assist


def main():
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")

    data_transform = transforms.Compose(
        [
         transforms.Resize(256),
         # transforms.Resize(448),
         transforms.CenterCrop(224),
         # transforms.CenterCrop(384),
         transforms.ToTensor(),

         # Image-Net
         transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
         ])
    # DataSet
    Data_num = 5
    # Selected Model from validation
    model_num = 132
    ori_path = '/media/hsy/1882C80C82C7EC76/ZJH/Cross-Validation-Dataset/Current_dataset/' + str(Data_num) + '/test.csv'
    test_images_path, test_images_label = read_csv_test(ori_path)
    # test_images_path, test_images_label = read_csv_test('./test.csv')
    # create record
    record = open('VOLO_dense_ConfusionMatrix_12_5_test.txt', mode='a')
    # record = open('Volo_dense_matrix_breast_200X_binary_record.txt', mode='a')
    record_2 = open('Wrong_picture_Volo_dense_12_5_test.txt', mode='a')
    # record_2 = open('Wrong_picture_Volo_dense_matrix_breast_200x_binary_record.txt', mode='a')
    record_score = open('Level2-3-score_12_5.txt', mode='a')

    # Create Excel
    excel_name = './12-5-Result.xlsx'
    if not os.path.exists(excel_name):
        Create_excel(excel_name)
    wb = openpyxl.load_workbook(excel_name)
    sheet = wb.active

    # 记录时间
    tm = time.strftime("%Y-%m-%d %X", time.localtime())
    record.write(tm + '\n')
    record_2.write(tm + '\n')
    if sheet.cell(row=Data_num+1, column=9).value is None:
        sheet.cell(row=Data_num+1, column=9).value = tm
    else:
        sheet.cell(row=Data_num + 11, column=9).value = tm
    # DATA-SET
    record.write('Data-Set:' + ori_path + '\n')
    record_2.write('Data-Set:' + ori_path + '\n')

    # Initialization
    k = 0
    count = 0
    # patient_num = 0
    # temp = []
    T1 = T2 = T3 = 0
    F1 = F2 = F3 = 0
    F12 = F13 = F21 = F23 = F31 = F32 = 0
    TP = TN = FP = FN = 0

    # create model
    # model = volo_d3(num_classes=3).to(device)
    model = volo_dense(num_classes=3).to(device)

    # load model weights
    # Load model path here.
    model_weight_path = "" + str(Data_num) + "/model-" + str(model_num) + ".pth"
    # model_weight_path = "./weights/model-" + str(model_num) + ".pth"

    model.load_state_dict(torch.load(model_weight_path, map_location=device))
    model.eval()
    record.write("Model name:model-" + str(model_num) + "\n")

    for file in test_images_path:
        # load image
        assert os.path.exists(file), "file: '{}' dose not exist.".format(file)
        img = Image.open(file)
        inc_factor = assist(img, device)
        # plt.imshow(img)
        # [N, C, H, W]
        img = data_transform(img)
        # expand batch dimension
        img = torch.unsqueeze(img, dim=0)

        with torch.no_grad():
            # predict class
            output = torch.squeeze(model(img.to(device))).cpu()
            predict = torch.softmax(output, dim=0)
            if abs(predict[1] - predict[2]) / predict[1] < 0.8:
                predict[2] += predict[2] * inc_factor
            elif abs(predict[1] - predict[2]) / predict[2] < 0.8:
                predict[2] += predict[2] * inc_factor
            else:
                pass

            predict_cla = torch.argmax(predict).numpy()
            record_score.write(f"Label:{str(test_images_label[k])}--------score:{str(predict)}\n")
            if predict_cla == test_images_label[k]:
                count += 1
            else:
                record_2.write(file + ' : ' + str(test_images_label[k]) + '------>' + str(predict_cla) + '\n')
                if test_images_label[k] == 2 or test_images_label[k] == 1:
                    record_2.write(f"Predicted result:{str(predict)}\n")


            # if predict_cla == 0 and test_images_label[k] == 0:
            #     TN += 1
            # elif predict_cla == 1 and test_images_label[k] == 0:
            #     FP += 1
            # elif predict_cla == 0 and test_images_label[k] == 1:
            #     FN += 1
            # else:
            #     TP += 1
            if predict_cla == 0 and test_images_label[k] == 0:
                T1 += 1
            elif predict_cla == 1 and test_images_label[k] == 0:
                F12 += 1
            elif predict_cla == 2 and test_images_label[k] == 0:
                F13 += 1
            elif predict_cla == 0 and test_images_label[k] == 1:
                F21 += 1
            elif predict_cla == 1 and test_images_label[k] == 1:
                T2 += 1
            elif predict_cla == 2 and test_images_label[k] == 1:
                F23 += 1
            elif predict_cla == 0 and test_images_label[k] == 2:
                F31 += 1
            elif predict_cla == 1 and test_images_label[k] == 2:
                F32 += 1
            elif predict_cla == 2 and test_images_label[k] == 2:
                T3 += 1
            else:
                pass

        print_res = "class: {}   true_class: {}".format(str(predict_cla),
                                                        test_images_label[k])
        # plt.title(print_res)
        print(print_res)
        k += 1
        # plt.show()

    print('\n' + str(count / len(test_images_label)))
    # print('\n' + 'Patient-Level:' + str((T1 + T2 + T3) / (len(test_images_label) // 12)))

    record.write('Accuracy:' + str(count / len(test_images_label)) + '\n')

    # record.write('Patient-Level-Accuracy:' + str((T1 + T2 + T3) / (len(test_images_label) // 12)) + '\n')
    record.write('-----------------------------------------------------\n')
    record.write('Confusion Matrix:\n')
    record.write('------ 1 ------- 2 ------ 3 ------\n')
    record.write('1----- ' + str(T1) + ' ------ ' + str(F12) + ' ------ ' + str(F13)+'\n')
    record.write('2----- ' + str(F21) + ' ------ ' + str(T2) + ' ----- ' + str(F23) + '\n')
    record.write('3----- ' + str(F31) + ' ------ ' + str(F32) + ' ------ ' + str(T3) + '\n')
    record.write('-----------------------------------------------------\n')
    # record.write('------ 0 ------- 1 \n')
    # record.write('0----- ' + str(TN) + ' ------ ' + str(FN) + '\n')
    # record.write('1----- ' + str(FP) + ' ------ ' + str(TP) + '\n')
    # record.write('-----------------------------------------------------\n')
    # record.write('------------Recall-------------------Precision-------------------F1-Score\n')

    # Recall = TP / (TP + FN)
    # Precision = TP / (TP + FP)
    # F1 = 2 * Recall * Precision / (Precision + Recall)
    # record.write('1----- ' + str(Recall) + ' ------ ' + str(Precision) + ' ------ ' + str(F1) + '\n')

    Recall_1 = T1 / (T1 + F12 + F13)
    Precision_1 = T1 / (T1 + F21 + F31)
    F1_S_1 = 2 * Recall_1 * Precision_1 / (Precision_1 + Recall_1)

    Recall_2 = T2 / (T2 + F21 + F23)
    Precision_2 = T2 / (T2 + F12 + F32)
    F1_S_2 = 2 * Recall_2 * Precision_2 / (Precision_2 + Recall_2)

    Recall_3 = T3 / (T3 + F31 + F32)
    Precision_3 = T3 / (T3 + F13 + F23)
    F1_S_3 = 2 * Recall_3 * Precision_3 / (Precision_3 + Recall_3)

    # record.write('1----- ' + str(Recall_1) + ' ------ ' + str(Precision_1) + ' ------ ' + str(F1_S_1) + '\n')
    # record.write('2----- ' + str(Recall_2) + ' ------ ' + str(Precision_2) + ' ------ ' + str(F1_S_2) + '\n')
    # record.write('3----- ' + str(Recall_3) + ' ------ ' + str(Precision_3) + ' ------ ' + str(F1_S_3) + '\n')

    # EXCEL
    accuracy = count / len(test_images_label)
    if sheet.cell(row=Data_num+1, column=2).value is None:
        sheet.cell(row=Data_num+1, column=2).value = accuracy
    else:
        sheet.cell(row=Data_num + 11, column=2).value = accuracy

    Average_Recall = (Recall_1 + Recall_2 + Recall_3) / 3
    Average_Precision = (Precision_1 + Precision_2 + Precision_3) / 3
    Average_F1 = (F1_S_1 + F1_S_2 + F1_S_3) / 3
    Average_FDR = 1 - Average_Recall
    Average_Error = 1 - accuracy
    # Average_Recall = Recall
    # Average_Precision = Precision
    # Average_F1 = F1

    if sheet.cell(row=Data_num+1, column=3).value is None:
        sheet.cell(row=Data_num+1, column=3).value = Average_Recall
    else:
        sheet.cell(row=Data_num + 11, column=3).value = Average_Recall

    if sheet.cell(row=Data_num+1, column=4).value is None:
        sheet.cell(row=Data_num+1, column=4).value = Average_Precision
    else:
        sheet.cell(row=Data_num + 11, column=4).value = Average_Precision

    if sheet.cell(row=Data_num+1, column=5).value is None:
        sheet.cell(row=Data_num+1, column=5).value = Average_F1
    else:
        sheet.cell(row=Data_num + 11, column=5).value = Average_F1

    if sheet.cell(row=Data_num+1, column=7).value is None:
        sheet.cell(row=Data_num+1, column=7).value = Average_Error
    else:
        sheet.cell(row=Data_num + 11, column=7).value = Average_Error

    if sheet.cell(row=Data_num+1, column=8).value is None:
        sheet.cell(row=Data_num+1, column=8).value = Average_FDR
    else:
        sheet.cell(row=Data_num + 11, column=8).value = Average_FDR
    #
    # Po = (T1 + T2 + T3) / (len(test_images_label))
    # Pe = ((T1+F12+F13)*(T1+F21+F31) + (F21+T2+F23)*(F12+T2+F32) + (F31+F32+T3)*(F13+F23+T3)) / ((len(test_images_label)) ** 2)
    # Kappa = (Po - Pe) / (1 - Pe)
    Po = (TP + TN) / len(test_images_label)
    Pe = ((TN+FN)*(TN+FP) + (TP+FP)*(TP+FN)) / ((len(test_images_label)) ** 2)
    Kappa = (Po - Pe) / (1 - Pe)

    if sheet.cell(row=Data_num+1, column=6).value is None:
        sheet.cell(row=Data_num+1, column=6).value = Kappa
    else:
        sheet.cell(row=Data_num + 11, column=6).value = Kappa

    record.write('Kappa:' + str(Kappa) + '\n\n')
    record_2.write('\n\n')
    record.close()
    record_2.close()
    wb.save(excel_name)


if __name__ == '__main__':
    main()
