import openpyxl
import csv

Res = {}
# with open('/media/hsy/1882C80C82C7EC76/ZJH/volo-main/T17芯片阵列排布表(周晨浩) - 分级(1).xlsx') as f:
#     f_csv = csv.reader(f)
#     for row in f_csv:
#         Res[row[0]] = row[1:]
#         # Res.append(row)
#         print(Res[row[0]])




def Create_excel(excel_name):
    wb = openpyxl.Workbook()
    # wb = openpyxl.load_workbook('./Vit_record.xlsx')
    sheet = wb.active
    sheet.title = 'Volo_dense'

    # wb.sheetnames

    # sheet = wb['ViT']
    sheet.cell(row=1, column=2).value = 'Accuracy'
    sheet.cell(row=1, column=3).value = 'Recall'
    sheet.cell(row=1, column=4).value = 'Precision'
    sheet.cell(row=1, column=5).value = 'F1-score'
    sheet.cell(row=1, column=6).value = 'Kappa'
    sheet.cell(row=1, column=7).value = 'Error-rate'
    sheet.cell(row=1, column=8).value = 'FDR'
    sheet.cell(row=1, column=9).value = 'Date'
    print(sheet.cell(row=2, column=2).value)
    # sheet = wb.active
    # sheet.title = 'Test'

    # wb.sheetnames

    wb.save(excel_name)

